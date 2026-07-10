"""
ingest.py — Reads a project-plan .xlsx export and produces a clean,
structured representation: leaf/parent task hierarchy, resolved baseline
dates, and merged free-text comments from both possible comment channels.

Every design choice here is backed by findings verified against the two
sample files (Project_Plan_B.xlsx, S2P_Project.xlsx) — see RAG_Methodology.md
and Data_Dictionary.md for the evidence behind each one.
"""

import openpyxl
import datetime
import re

# Columns confirmed structurally dead (#UNPARSEABLE on every row in both
# sample files) — dropped unconditionally, never treated as "missing data".
DEAD_COLUMNS = {
    "No.of days Until Today",
    "No.of days",
    "Target start date to Today",
    "Project Category",
}

# Columns confirmed to be empty/ghost duplicates in at least one sample file
# (e.g. Plan B's 'Start'/'Finish'/'Baseline Start Date'/'Baseline End Date').
# We don't hard-drop these globally since a different export might populate
# them — instead ingest() checks fill-rate at read time (see _pick_baseline_cols).
KNOWN_GHOST_COLUMNS = {"Start", "Finish", "Baseline Start Date", "Baseline End Date"}


def _col_index(headers):
    return {h: i for i, h in enumerate(headers) if h is not None}


def _fill_rate(rows, i):
    if i is None:
        return 0
    return sum(1 for r in rows if r[i] not in (None, "")) / max(len(rows), 1)


def _pick_baseline_cols(idx, rows):
    """
    Verified finding: Plan B has two candidate baseline-date column sets
    ('Baseline Start'/'Baseline Finish', almost entirely blank — 1/383 rows —
    and 'Baseline Start2'/'Baseline Finish2', populated on 148/383 rows).
    S2P only has the plain set, populated on 272/493 rows.
    Rule: prefer whichever set has a higher fill rate; if both exist, take
    the '2' variant per-row when present, else fall back to the plain one.
    """
    has_v2 = "Baseline Start2" in idx and "Baseline Finish2" in idx
    plain_i, plain_f = idx.get("Baseline Start"), idx.get("Baseline Finish")
    v2_i, v2_f = idx.get("Baseline Start2"), idx.get("Baseline Finish2")

    if has_v2 and _fill_rate(rows, v2_i) >= _fill_rate(rows, plain_i):
        return v2_i, v2_f, plain_i, plain_f
    return plain_i, plain_f, v2_i, v2_f


def _parse_predecessors(raw):
    """Returns (list_of_refs, has_broken_ref). Verified format: 'N', 'NFS +7d',
    comma-separated, and occasionally a broken '#REF' link (found: 4 in Plan B)."""
    if not raw:
        return [], False
    broken = False
    refs = []
    for tok in str(raw).split(","):
        tok = tok.strip()
        if "#REF" in tok:
            broken = True
            continue
        m = re.match(r"^(\d+)", tok)
        if m:
            refs.append(int(m.group(1)))
    return refs, broken


def _find_main_sheet(wb):
    for name in wb.sheetnames:
        if name not in ("Comments", "Summary"):
            return name
    raise ValueError("Could not find a main task sheet (expected one sheet besides Comments/Summary)")


def _read_comments_sheet(wb):
    """Verified: this sheet has NO header row (an earlier pass of ours
    mistakenly treated its first entry as a header). Format per row:
    (row_ref like 'Row 292', comment text, author, timestamp)."""
    if "Comments" not in wb.sheetnames:
        return []
    ws = wb["Comments"]
    out = []
    for r in ws.iter_rows(min_row=1, values_only=True):
        if not any(r):
            continue
        row_ref, text, author, ts = (list(r) + [None, None, None, None])[:4]
        if not row_ref or not text:
            continue
        m = re.match(r"Row\s*(\d+)", str(row_ref))
        if not m:
            continue
        out.append({
            "excel_row": int(m.group(1)),
            "text": str(text).strip(),
            "author": author,
            "timestamp": str(ts) if ts else None,
            "source": "comments_sheet",
        })
    return out


def _read_summary_sheet(wb):
    if "Summary" not in wb.sheetnames:
        return {}
    ws = wb["Summary"]
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] is not None:
            d[str(row[0]).strip()] = row[1] if len(row) > 1 else None
    return d


def load_project(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _find_main_sheet(wb)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = _col_index(headers)
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    bs_i, bf_i, bs_alt, bf_alt = _pick_baseline_cols(idx, raw_rows)

    anc_i = idx.get("Ancestors")
    if anc_i is None:
        raise ValueError("Expected an 'Ancestors' column to determine task hierarchy")

    # Leaf detection: a row is a parent if the NEXT row goes one level deeper.
    n = len(raw_rows)
    is_leaf = [True] * n
    for i in range(n - 1):
        a_cur, a_next = raw_rows[i][anc_i], raw_rows[i + 1][anc_i]
        if a_cur is not None and a_next is not None and a_next > a_cur:
            is_leaf[i] = False

    def g(row, name, default=None):
        i = idx.get(name)
        if i is None:
            return default
        v = row[i]
        return v if v not in ("", None) else default

    tasks = []
    for i, row in enumerate(raw_rows):
        excel_row = i + 2  # header is row 1
        preds, broken_pred = _parse_predecessors(g(row, "Predecessors"))

        baseline_start = row[bs_i] if bs_i is not None else None
        baseline_finish = row[bf_i] if bf_i is not None else None
        if baseline_start is None and bs_alt is not None:
            baseline_start = row[bs_alt]
        if baseline_finish is None and bf_alt is not None:
            baseline_finish = row[bf_alt]

        task = {
            "excel_row": excel_row,
            "ancestors": row[anc_i],
            "is_leaf": is_leaf[i],
            "task_name": g(row, "Task Name"),
            "status": g(row, "Status"),
            "pct_complete": g(row, "% Complete"),
            "start_date": g(row, "Start Date"),
            "end_date": g(row, "End Date"),
            "baseline_start": baseline_start,
            "baseline_finish": baseline_finish,
            "total_float": g(row, "Total Float"),
            "critical": bool(g(row, "Critical ?", False)),
            "predecessors": preds,
            "broken_predecessor_ref": broken_pred,
            "phase_milestone": g(row, "Phase/Milestone"),
            "priority": g(row, "Priority"),
            "at_risk": bool(g(row, "At Risk?", False)),
            "schedule_health_src": g(row, "Schedule Health"),
            "rag_src": g(row, "RAG"),
            "owner": g(row, "Owner"),
            "assigned_to": g(row, "Assigned To"),
            "row_comment": g(row, "Comments"),
        }
        tasks.append(task)

    # Merge both comment channels: the row-level 'Comments' column, and the
    # separate Comments sheet (matched back to a task by its excel row number).
    comments = []
    for t in tasks:
        if t["row_comment"]:
            comments.append({
                "excel_row": t["excel_row"],
                "text": t["row_comment"],
                "author": None,
                "timestamp": None,
                "source": "row_column",
            })
    comments.extend(_read_comments_sheet(wb))

    by_row = {t["excel_row"]: t for t in tasks}
    for c in comments:
        c["task"] = by_row.get(c["excel_row"])

    project_name = tasks[0]["task_name"] if tasks else None

    return {
        "file_path": path,
        "sheet_name": sheet_name,
        "project_name": project_name,
        "tasks": tasks,
        "comments": comments,
        "summary_sheet": _read_summary_sheet(wb),
        "dropped_columns": sorted(DEAD_COLUMNS & set(idx.keys())),
    }


if __name__ == "__main__":
    import sys
    import json

    for p in ("/mnt/user-data/uploads/Project_Plan_B.xlsx", "/mnt/user-data/uploads/S2P_Project.xlsx"):
        proj = load_project(p)
        leaf_count = sum(1 for t in proj["tasks"] if t["is_leaf"])
        print(f"{p}")
        print(f"  project_name={proj['project_name']!r}")
        print(f"  total_tasks={len(proj['tasks'])} leaf_tasks={leaf_count}")
        print(f"  comments merged={len(proj['comments'])} (with task match: {sum(1 for c in proj['comments'] if c['task'])})")
        print(f"  dropped dead columns: {proj['dropped_columns']}")
        print()
